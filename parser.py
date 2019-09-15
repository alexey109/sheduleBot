#!/usr/bin/env python
# -*- coding: UTF-8 -*-


from openpyxl import load_workbook
import re


def safeGet(lst, idx, default):
    """ Helpful function if you want to get list value by idx.
    
    :param lst: list
    :type lst: list
    :param idx: value index
    :type idx: int
    :param default: return it if index out of list 
    :type default: any
    :return: 'default' argument 
    :rtype: any
    """
    try:
        return lst[idx]
    except:
        return default


def getDayNumber(sval):
    days = {
        u'понедельник': 0,
        u'вторник': 1,
        u'среда': 2,
        u'четверг': 3,
        u'пятница': 4,
        u'суббота': 5,
        u'воскресение': 6
    }
    return days.get(sval.lower(), 0)


class Parser:
    """
    Retrieves schedule from excel document.
    """

    def openDoc(self, doc_name):
        """ Open excel-document.
        
        :param doc_name: document location with name
        :type doc_name: str
        :return: document object
        :rtype: object
        """
        result = True
        try:
            self.wb = load_workbook(filename=doc_name)
        except Exception as e:
            print doc_name
            print e
            result = False
        return result

    def splitBody(self, text):
        """ In documents main lesson information contains in one cell.
        Thus function parse text from that cell.
        
        :param text: text from cell
        :type text: str
        :return: lessons information
        :rtype: list
        """
        text = text.replace(u'\n', ' ')
        text = text.replace(u'н.', '') + '.'  # Bag fix :(

        ST_BODY = 0
        ST_PARAMS = 1
        ST_WEEK = 2

        teacher_kwd = [u'асс', u'доц', u'проф', u'ст.', u'ст ',
                       u'преп']

        # вычленяем шаблон "кр 10" (кроме 10 недели)
        exclude_week = None
        match = re.search(u'кр\.?\s?[1-9][0-9]?', text)
        if match:
            exclude_week = match.group(0)
            text = text.replace(exclude_week, '')

        rules = {
            ST_BODY: [
                {'expr': u'[А-Яа-я\.\s\/]', 'link': ST_BODY,
                 'ext': False, 'new_param': False},
                {'expr': '[0-9I]', 'link': ST_WEEK, 'ext': True,
                 'new_param': False},
                {'expr': '\(', 'link': ST_PARAMS, 'ext': False,
                 'new_param': True},
            ],
            ST_PARAMS: [
                {'expr': '[^\)]', 'link': ST_PARAMS, 'ext': False,
                 'new_param': False},
                {'expr': '\)', 'link': ST_BODY, 'ext': False,
                 'new_param': False},
            ],
            ST_WEEK: [
                {'expr': '[0-9I,.\-\s]', 'link': ST_WEEK, 'ext': False,
                 'new_param': False},
                {'expr': '[^0-9I,\-\s]', 'link': ST_BODY,
                 'ext': False, 'new_param': False},
            ]
        }

        status = ST_BODY
        i = 0
        body = {}
        params = []
        name = week = teacher = ''
        for idx, l in enumerate(text):
            for rule in rules[status]:
                if re.search(rule['expr'], l):
                    status = rule['link']
                    if (rule['ext'] and name) or (idx == len(text) - 1):
                        for word in teacher_kwd:
                            pos = name.find(word)
                            if pos > 0:
                                teacher = name[pos:]
                                name = name[:pos]
                                break

                        try:
                            if body[i-1]['name'].strip().endswith(u'с'):
                                week = week.strip() + '-17'
                                i -= 1
                        except:
                            pass
                        try:
                            if body[i-1]['name'].strip().endswith(u'кр'):
                                params.append(u'кр {}'.format(week))
                                week = ''
                                i -= 1
                        except:
                            pass
                        if len(name.strip()) < 2:
                            continue
                        body[i] = {}
                        body[i]['name'] = name.strip()
                        body[i]['teacher'] = teacher.strip()
                        body[i]['params'] = params
                        body[i]['week'] = exclude_week if exclude_week else week.strip()
                        name = week = teacher = ''
                        params = []
                        i += 1
                    if rule['new_param']:
                        params.append('')
                    break

            if l in ['(', ')']:
                continue

            if status == ST_BODY:
                name += l
            elif status == ST_PARAMS:
                try:
                    params[-1] += l
                except:
                    params.append(l)  #  Fix if no ')' in the string
            elif status == ST_WEEK:
                week += l

        return body

    def getGroupSchdl(self, schedule, sheet):
        """ Look up for one column and get schedule for one group.
        
        :param schedule: full schedule
        :type schedule: dict
        :param sheet: excel sheet object
        :type sheet: object
        :return: schedule with new added group
        :rtype: dict
        """

        def isEventEqual(event1, event2):
            """ Check two lesson for equation.
            
            :return: Equal or not
            :rtype: bool
            """
            equal = False
            if event1['day'] == event2['day'] \
                    and event1['numb'] == event2['numb'] \
                    and event1['room'] == event2['room'] \
                    and event1['name'] == event2['name'] \
                    and event1['teacher'] == event2['teacher']:
                try:
                    params = True
                    for i in range(0, len(event1['params'])):
                        if params and (
                                    event1['params'][i] !=
                                    event2['params'][i]):
                            params = False
                except:
                    params = False

                if params and 'I' in event1['week'] \
                        and 'I' in event2['week']:
                    equal = True
            return equal

        def checkSkipWords(name):
            """ Check if full day event on a day.
            
            :param name: cell value
            :type name: str
            :return: is full day event or not
            :rtype: bool
            """
            keywords = [u'воен', u'производств.*практ', u'кафедр',
                        u'самостоятел', u'занят',
                        u'день', u'адрес', u'пироговск']
            find = False
            for keyword in keywords:
                if re.search(keyword, name.lower()):
                    find = True
                    break
            return find

        lessons_numbers = {}
        days_numbers = {}
        gotit = False
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=4):
            for cell in row:
                if cell.value and unicode(cell.value).find(u'пары') > 0:
                    for i in range(cell.row+1, 77):
                        # fill lesson number
                        cur_cell = sheet.cell(row=i, column=cell.col_idx)
                        if cur_cell.value:
                            lessons_numbers[i] = {
                                'numb': int(cur_cell.value),
                                'type': 'I'
                            }
                        else:
                            lessons_numbers[i] = {
                                'numb': lessons_numbers[i-1]['numb'],
                                'type': 'II'
                            }
                        # fill days number
                        cur_cell = sheet.cell(row=i, column=cell.col_idx-1)
                        if cur_cell.value:
                            days_numbers[i] = getDayNumber(cur_cell.value)
                        else:
                            days_numbers[i] = days_numbers[i - 1]
                    gotit = True
                    break
            if gotit:
                break

        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=3):
            for cell in row:
                lector_flag = False
                if not cell.value:
                    continue
                try:
                    match = re.search(
                        u"[А-Яа-я]{4}[А-Яа-я]?-[0-9]{2}-[0-9]{2}",
                        cell.value)
                    if not match:
                        continue
                except:
                    continue
                try:
                    j = cell.col_idx
                    row_start = cell.row + 2
                    group = cell.value.lower()
                    lections = []
                    for i in range(row_start, 77):
                        content = sheet.cell(row=i, column=j).value
                        if not content:
                            continue

                        etype = sheet.cell(row=i, column=j + 1).value
                        lectors = sheet.cell(row=i, column=j + 2).value
                        rooms = sheet.cell(row=i, column=j + 3).value

                        cell_day = days_numbers[i]
                        cell_numb = lessons_numbers[i]['numb']
                        cell_week = lessons_numbers[i]['type']
                        info = self.splitBody(content)

                        if len(info) > 1:
                            lectors = re.split('\n|\s{3,10}', lectors) if lectors else []
                            etype = re.split('\n|\s{3,10}', etype) if etype else ''
                            rooms = re.split('\n|\s{3,10}', unicode(rooms)) if rooms else []
                        else:
                            lectors = [lectors] if lectors else ['']
                            etype = [etype] if etype else ['']
                            rooms = [unicode(rooms).replace('\n', ' ')] if rooms else ['']
                        for i in range(0, len(info)):
                            if len(info[i]['name']) < 2 or checkSkipWords(info[i]['name']):
                                continue
                            event_type = safeGet(etype, i, '')
                            lector = safeGet(lectors, i, '-')
                            room = safeGet(rooms, i, '-')

                            week = info[i]['week'] if info[i]['week'] else cell_week
                            if '-' in week:
                                week = ','.join([str(iw) for iw in range(int(week.split('-')[0]), int(week.split('-')[1])+1, 2)])
                            if u'кр' in week:
                                number = int(re.sub('[^0-1]', '', week))
                                week = ','.join([str(iw) for iw in range(1, 18) if iw != number])

                            event = {
                                'day': cell_day,
                                'numb': cell_numb,
                                'room': room[:20],
                                'week': week[:50],
                                'name': (info[i]['name'] + ' ' + event_type.replace('\n', ''))[:100],
                                'teacher': lector[:60],
                                'params': info[i]['params']
                            }
                            append_flag = True
                            for l in lections:
                                if isEventEqual(l, event):
                                    l['week'] = ''
                                    append_flag = False

                            if append_flag:
                                lections.append(event)

                    print(group)
                    schedule[group] = lections
                except Exception as e:
                    print 'Exception when parse group ' + group + ' schedule: ' + str(e)
                    continue

        return schedule

    def getSchedule(self, document):
        """ Retrieves groups schedule from excel document.
        
        :param document: document location with name
        :type document: str
        :return: schedule type (lections/exams/zachet)
        :return: groups schedule
        :rtype: dict
        """
        if not self.openDoc(document):
            raise Exception("document doesn't open")

        schedule = {}
        schdl_type = 'lections'
        exam = False
        for sheet in self.wb:

            for row in sheet.iter_rows(min_row=1,
                                       max_col=sheet.max_column,
                                       max_row=3):
                for cell in row:
                    try:
                        group_name = cell.value
                        if u'экзамен' in group_name.lower():
                            exam = True
                            schdl_type = 'exams'

                        if u'зачет' in group_name.lower():
                            exam = True
                            schdl_type = 'zachet'
                    except:
                        pass

            if not exam:
                schedule = self.getGroupSchdl(schedule, sheet)
                break

        return schdl_type, schedule
